"""Extract structured metadata from DDL, SQL, requirement docs, and code tables.

Produces schema.md, metrics.md, code_tables.md, sample_sql.md under meta/.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import sqlglot
from sqlglot import exp

from services.llm_client import chat_completion_json

META_DIR = Path(__file__).resolve().parent.parent / "meta"

try:
    import openpyxl  # type: ignore
except Exception:  # noqa: BLE001
    openpyxl = None


# ---------------------------------------------------------------------------
# DDL -> schema.md
# ---------------------------------------------------------------------------

def parse_ddl(ddl_text: str) -> list[dict]:
    """Parse DDL statements and return a list of table definitions."""
    tables: list[dict] = []
    try:
        stmts = sqlglot.parse(ddl_text, read="hive", error_level=sqlglot.ErrorLevel.IGNORE)
    except Exception:
        stmts = []

    for stmt in stmts:
        if not isinstance(stmt, exp.Create):
            continue
        table_name = stmt.find(exp.Table)
        if table_name is None:
            continue
        tbl: dict[str, Any] = {
            "name": table_name.name,
            "schema": table_name.db or "",
            "columns": [],
            "comment": "",
        }
        schema_node = stmt.find(exp.Schema)
        if schema_node:
            for col_def in schema_node.find_all(exp.ColumnDef):
                col: dict[str, str] = {
                    "name": col_def.name,
                    "type": col_def.args.get("kind", exp.DataType(this=exp.DataType.Type.VARCHAR)).sql(),
                    "comment": "",
                }
                for prop in col_def.find_all(exp.ColumnConstraint):
                    comment_node = prop.find(exp.SchemaCommentColumnConstraint)
                    if comment_node:
                        col["comment"] = comment_node.this.this if comment_node.this else ""
                tbl["columns"].append(col)
        table_comment = stmt.find(exp.SchemaCommentColumnConstraint)
        if table_comment and table_comment.this:
            tbl["comment"] = table_comment.this.this
        tables.append(tbl)
    return tables


def parse_sql_for_joins(sql_text: str) -> list[dict]:
    """Parse sample SQL to extract JOIN relationships and filter patterns."""
    joins: list[dict] = []
    cleaned_sql = _extract_sql_chunks(sql_text)
    try:
        stmts = sqlglot.parse(cleaned_sql, read="hive", error_level=sqlglot.ErrorLevel.IGNORE)
    except Exception:
        return joins

    for stmt in stmts:
        if stmt is None:
            continue
        for join_node in stmt.find_all(exp.Join):
            table_node = join_node.find(exp.Table)
            on_node = join_node.args.get("on")
            if table_node:
                joins.append({
                    "table": table_node.name,
                    "alias": table_node.alias or "",
                    "join_type": join_node.args.get("side", ""),
                    "on_condition": on_node.sql() if on_node else "",
                })
    return joins


def _extract_sql_chunks(text: str) -> str:
    """Best-effort extraction of SQL statements from markdown-ish text."""
    if not text:
        return ""
    fenced = re.findall(r"```sql\\s*([\\s\\S]*?)```", text, flags=re.IGNORECASE)
    if fenced:
        return "\n\n".join(s.strip() for s in fenced if s.strip())
    lines: list[str] = []
    started = False
    for line in text.splitlines():
        t = line.strip()
        if not started and re.search(r"\\bSELECT\\b", t, flags=re.IGNORECASE):
            started = True
        if not started:
            continue
        if t.startswith("#") or t.startswith("##") or t.startswith("-"):
            continue
        lines.append(line)
    return "\n".join(lines)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "")


def _read_xlsx_rows(path: Path, *, header_row: int = 1, max_rows: int = 50000) -> list[dict[str, Any]]:
    """Read an xlsx sheet into list-of-dict rows.

    Notes:
    - This is a best-effort parser. If column names are unknown, we still keep raw columns.
    - We intentionally cap rows to avoid oversized meta files.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed; please `pip install -r requirements.txt`")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        headers = [_normalize_header(v) or f"COL_{i+1}" for i, v in enumerate(row)]
        break

    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=1):
        if idx > max_rows:
            break
        if row is None:
            continue
        rec: dict[str, Any] = {}
        empty = True
        for i, v in enumerate(row):
            key = headers[i] if i < len(headers) else f"COL_{i+1}"
            rec[key] = v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(rec)
    return rows


def parse_data_dictionary_xlsx(path: Path) -> list[dict[str, Any]]:
    """Parse '数据字典' xlsx to table definitions compatible with _build_schema_md."""
    rows = _read_xlsx_rows(path)
    # Common column guesses
    candidates_table = ("表名", "TABLE_NAME", "table_name", "TABLE", "模型名称")
    candidates_col = ("字段名", "字段名称", "COLUMN_NAME", "col_name", "字段", "模型字段英文名称")
    candidates_type = ("类型", "字段类型", "DATA_TYPE", "type", "字段类型")
    candidates_comment = ("含义", "字段含义", "注释", "COMMENT", "remark", "备注", "模型字段中文名称")
    candidates_table_comment = ("表描述", "模型描述", "comment", "描述")

    def pick_key(rec: dict[str, Any], keys: tuple[str, ...]) -> str | None:
        for k in rec.keys():
            if k in keys:
                return k
        # fuzzy contains
        for k in rec.keys():
            for kk in keys:
                if kk and kk.lower() in str(k).lower():
                    return k
        return None

    tables: dict[str, dict[str, Any]] = {}
    for r in rows:
        k_table = pick_key(r, candidates_table)
        k_col = pick_key(r, candidates_col)
        if not k_table or not k_col:
            continue
        table_name = str(r.get(k_table) or "").strip()
        col_name = str(r.get(k_col) or "").strip()
        if not table_name or not col_name:
            continue
        k_type = pick_key(r, candidates_type)
        k_comment = pick_key(r, candidates_comment)

        t = tables.setdefault(
            table_name,
            {"name": table_name, "schema": "", "columns": [], "comment": ""},
        )
        k_tbl_c = pick_key(r, candidates_table_comment)
        if k_tbl_c and not t.get("comment") and r.get(k_tbl_c):
            t["comment"] = str(r.get(k_tbl_c) or "").strip()
        t["columns"].append(
            {
                "name": col_name,
                "type": str(r.get(k_type) or "").strip() if k_type else "",
                "comment": str(r.get(k_comment) or "").strip() if k_comment else "",
            }
        )
    return list(tables.values())


def parse_metrics_kb_xlsx(path: Path) -> list[dict[str, Any]]:
    """Parse '指标口径知识库' xlsx to metric dicts compatible with _build_metrics_md."""
    rows = _read_xlsx_rows(path)
    # Common column guesses
    candidates_name = ("指标名称", "指标名", "name", "NAME", "字段名称", "字段名")
    candidates_en = ("英文名", "english_name", "EN_NAME")
    candidates_def = ("业务含义", "含义", "definition", "DEF")
    candidates_calc = ("计算逻辑", "口径", "calculation", "SQL", "公式", "技术口径")
    candidates_table = ("来源表", "source_table", "TABLE", "表名")
    candidates_gran = ("统计粒度", "粒度", "granularity")
    candidates_cons = ("限制条件", "约束", "constraints")

    def pick_key(keys: tuple[str, ...]) -> str | None:
        if not rows:
            return None
        # use header from first row
        header_keys = list(rows[0].keys())
        for hk in header_keys:
            if hk in keys:
                return hk
        for hk in header_keys:
            for kk in keys:
                if kk and kk.lower() in str(hk).lower():
                    return hk
        return None

    k_name = pick_key(candidates_name)
    if not k_name:
        return []
    k_en = pick_key(candidates_en)
    k_def = pick_key(candidates_def)
    k_calc = pick_key(candidates_calc)
    k_table = pick_key(candidates_table)
    k_gran = pick_key(candidates_gran)
    k_cons = pick_key(candidates_cons)

    metrics: list[dict[str, Any]] = []
    for r in rows:
        name = str(r.get(k_name) or "").strip()
        if not name:
            continue
        cons_val = str(r.get(k_cons) or "").strip() if k_cons else ""
        metrics.append(
            {
                "name": name,
                "english_name": str(r.get(k_en) or "").strip() if k_en else "",
                "definition": str(r.get(k_def) or "").strip() if k_def else "",
                "calculation": str(r.get(k_calc) or "").strip() if k_calc else "",
                "source_table": str(r.get(k_table) or "").strip() if k_table else "",
                "granularity": str(r.get(k_gran) or "").strip() if k_gran else "",
                "constraints": [c.strip() for c in re.split(r"[，,;；\n]+", cons_val) if c.strip()] if cons_val else [],
            }
        )
    return metrics


def _build_schema_md(tables: list[dict], joins: list[dict]) -> str:
    """Render tables and joins into schema.md markdown."""
    lines: list[str] = ["# 数据字典 (schema)\n"]
    for tbl in tables:
        comment = f"\n{tbl['comment']}" if tbl.get("comment") else ""
        prefix = f"{tbl['schema']}." if tbl.get("schema") else ""
        lines.append(f"## {prefix}{tbl['name']}{comment}\n")
        if tbl["columns"]:
            lines.append("| 字段名 | 类型 | 含义 |")
            lines.append("|--------|------|------|")
            for c in tbl["columns"]:
                lines.append(f"| {c['name']} | {c['type']} | {c['comment']} |")
        lines.append("")

    if joins:
        lines.append("## 表关联关系\n")
        lines.append("| 关联表 | 别名 | JOIN 类型 | ON 条件 |")
        lines.append("|--------|------|-----------|---------|")
        for j in joins:
            lines.append(f"| {j['table']} | {j['alias']} | {j['join_type']} | {j['on_condition']} |")
        lines.append("")
    return "\n".join(lines)


def _build_sample_sql_md(raw_sql: str) -> str:
    lines = ["# 样例 SQL 库\n"]
    lines.append("以下为参考样例 SQL，可用于学习 JOIN 模式和编码风格。\n")
    lines.append("```sql")
    lines.append(raw_sql.strip())
    lines.append("```\n")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Requirement doc / code table helpers
# ---------------------------------------------------------------------------

def parse_code_table_md(text: str) -> str:
    """Pass-through: code table markdown is already structured."""
    lines = ["# 码表数据\n"]
    lines.append(text.strip())
    lines.append("")
    return "\n".join(lines)


async def extract_metrics_from_text(requirement_text: str) -> str:
    """Use LLM to extract metric definitions from a requirement document."""
    system = (
        "你是一个数仓指标口径分析专家。请从以下需求文档中提取所有指标定义，"
        "输出 JSON 数组，每个元素包含：name(指标中文名), english_name(建议英文名), "
        "definition(业务含义), calculation(计算逻辑/SQL片段), source_table(来源表), "
        "granularity(统计粒度), constraints(限制条件数组)。"
        "仅基于文档内容提取，不要编造。"
    )
    try:
        result = await chat_completion_json(system, requirement_text)
        metrics = result if isinstance(result, list) else result.get("metrics", [])
    except Exception:
        metrics = _extract_metrics_regex(requirement_text)

    return _build_metrics_md(metrics)


def _extract_metrics_regex(text: str) -> list[dict]:
    """Fallback: extract metrics using regex patterns."""
    metrics: list[dict] = []
    patterns = [
        r"(?:指标|字段)[名称]*[：:]\s*(.+?)(?:\n|$)",
        r"COUNT\s*\(\s*DISTINCT\s+(\w+)\s*\)",
    ]
    names_seen: set[str] = set()
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            name = m.group(1).strip()
            if name not in names_seen:
                names_seen.add(name)
                metrics.append({"name": name, "definition": "", "calculation": m.group(0)})
    return metrics


def _build_metrics_md(metrics: list[dict]) -> str:
    lines = ["# 指标口径库 (metrics)\n"]
    if not metrics:
        lines.append("_暂无指标数据，请上传需求文档或手动编辑。_\n")
        return "\n".join(lines)

    lines.append("| 指标名称 | 英文名 | 业务含义 | 计算逻辑 | 来源表 | 统计粒度 | 限制条件 |")
    lines.append("|----------|--------|----------|----------|--------|----------|----------|")
    for m in metrics:
        constraints = ", ".join(m.get("constraints", [])) if m.get("constraints") else ""
        lines.append(
            f"| {m.get('name', '')} "
            f"| {m.get('english_name', '')} "
            f"| {m.get('definition', '')} "
            f"| {m.get('calculation', '')} "
            f"| {m.get('source_table', '')} "
            f"| {m.get('granularity', '')} "
            f"| {constraints} |"
        )
    lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# High-level orchestrator
# ---------------------------------------------------------------------------

async def extract_and_save(
    *,
    ddl_text: str | None = None,
    sample_sql_text: str | None = None,
    requirement_text: str | None = None,
    code_table_text: str | None = None,
    data_dictionary_xlsx_paths: list[Path] | None = None,
    metrics_kb_xlsx_paths: list[Path] | None = None,
) -> dict[str, str]:
    """Run full extraction pipeline and persist to meta/ directory."""
    META_DIR.mkdir(parents=True, exist_ok=True)
    saved: dict[str, str] = {}

    tables: list[dict] = []
    joins: list[dict] = []

    if ddl_text:
        tables = parse_ddl(ddl_text)
    if data_dictionary_xlsx_paths:
        for p in data_dictionary_xlsx_paths:
            try:
                tables.extend(parse_data_dictionary_xlsx(p))
            except Exception:
                continue
    if sample_sql_text:
        joins = parse_sql_for_joins(sample_sql_text)

    if tables or joins:
        schema_md = _build_schema_md(tables, joins)
        (META_DIR / "schema.md").write_text(schema_md, encoding="utf-8")
        saved["schema.md"] = schema_md

    if sample_sql_text:
        sql_md = _build_sample_sql_md(sample_sql_text)
        (META_DIR / "sample_sql.md").write_text(sql_md, encoding="utf-8")
        saved["sample_sql.md"] = sql_md

    if metrics_kb_xlsx_paths:
        all_metrics: list[dict[str, Any]] = []
        for p in metrics_kb_xlsx_paths:
            try:
                all_metrics.extend(parse_metrics_kb_xlsx(p))
            except Exception:
                continue
        metrics_md = _build_metrics_md(all_metrics)
        (META_DIR / "metrics.md").write_text(metrics_md, encoding="utf-8")
        saved["metrics.md"] = metrics_md
    elif requirement_text:
        metrics_md = await extract_metrics_from_text(requirement_text)
        (META_DIR / "metrics.md").write_text(metrics_md, encoding="utf-8")
        saved["metrics.md"] = metrics_md

    if code_table_text:
        ct_md = parse_code_table_md(code_table_text)
        (META_DIR / "code_tables.md").write_text(ct_md, encoding="utf-8")
        saved["code_tables.md"] = ct_md

    return saved


def load_meta_file(name: str) -> str | None:
    """Load a metadata file from meta/ directory."""
    p = META_DIR / name
    if p.exists():
        return p.read_text(encoding="utf-8")
    return None


def list_meta_files() -> dict[str, bool]:
    """Return which meta files exist."""
    names = ["schema.md", "metrics.md", "code_tables.md", "sample_sql.md"]
    return {n: (META_DIR / n).exists() for n in names}
