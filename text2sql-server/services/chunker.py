"""Document chunker: split files into retrieval-friendly chunks.

Supported formats: Markdown (.md), Excel (.xlsx), plain text (.txt).
Each chunk carries metadata (source file, type, section) for traceability.
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any

try:
    import openpyxl
except Exception:
    openpyxl = None

MAX_CHUNK_CHARS = 500
OVERLAP_CHARS = 50


def _make_id() -> str:
    return uuid.uuid4().hex[:12]


def _sliding_window(text: str, *, max_chars: int = MAX_CHUNK_CHARS, overlap: int = OVERLAP_CHARS) -> list[str]:
    """Split long text into overlapping windows."""
    if len(text) <= max_chars:
        return [text]
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = start + max_chars
        chunks.append(text[start:end])
        start = end - overlap
    return chunks


def chunk_markdown(text: str, *, source_file: str = "") -> list[dict[str, Any]]:
    """Split markdown by ## headings, then apply sliding window if needed."""
    sections: list[tuple[str, str]] = []
    current_title = ""
    current_lines: list[str] = []

    for line in text.splitlines():
        if re.match(r"^#{1,3}\s+", line):
            if current_lines:
                sections.append((current_title, "\n".join(current_lines).strip()))
            current_title = line.lstrip("#").strip()
            current_lines = [line]
        else:
            current_lines.append(line)

    if current_lines:
        sections.append((current_title, "\n".join(current_lines).strip()))

    chunks: list[dict[str, Any]] = []
    for title, body in sections:
        if not body.strip():
            continue
        for window in _sliding_window(body):
            chunks.append({
                "id": _make_id(),
                "text": window,
                "metadata": {
                    "source_file": source_file,
                    "chunk_type": "markdown_section",
                    "section_title": title,
                },
            })
    return chunks


def chunk_excel(path: Path, *, source_file: str = "") -> list[dict[str, Any]]:
    """Split each sheet's rows into chunks, grouped by a key column when possible."""
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    chunks: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers: list[str] = []
        rows_data: list[list[Any]] = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(v).strip() if v else f"COL_{j}" for j, v in enumerate(row)]
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rows_data.append(list(row))

        if not headers or not rows_data:
            continue

        group_col_idx = _find_group_column(headers)
        if group_col_idx is not None:
            groups: dict[str, list[list[Any]]] = {}
            for row in rows_data:
                key = str(row[group_col_idx] or "").strip() or "未知"
                groups.setdefault(key, []).append(row)
            for group_key, group_rows in groups.items():
                text = _rows_to_text(headers, group_rows)
                for window in _sliding_window(text):
                    chunks.append({
                        "id": _make_id(),
                        "text": window,
                        "metadata": {
                            "source_file": source_file,
                            "chunk_type": "excel_row",
                            "section_title": f"{sheet_name} / {group_key}",
                        },
                    })
        else:
            batch_size = 10
            for start in range(0, len(rows_data), batch_size):
                batch = rows_data[start : start + batch_size]
                text = _rows_to_text(headers, batch)
                for window in _sliding_window(text):
                    chunks.append({
                        "id": _make_id(),
                        "text": window,
                        "metadata": {
                            "source_file": source_file,
                            "chunk_type": "excel_row",
                            "section_title": sheet_name,
                        },
                    })
    return chunks


def chunk_text(text: str, *, source_file: str = "") -> list[dict[str, Any]]:
    """Split plain text by paragraphs (double newlines)."""
    paragraphs = re.split(r"\n\s*\n", text)
    chunks: list[dict[str, Any]] = []
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        for window in _sliding_window(para):
            chunks.append({
                "id": _make_id(),
                "text": window,
                "metadata": {
                    "source_file": source_file,
                    "chunk_type": "text_paragraph",
                    "section_title": "",
                },
            })
    return chunks


def chunk_file(path: Path) -> list[dict[str, Any]]:
    """Auto-detect file type and produce chunks."""
    suffix = path.suffix.lower()
    source_file = path.name

    if suffix in (".md", ".markdown"):
        text = path.read_text(encoding="utf-8")
        return chunk_markdown(text, source_file=source_file)
    elif suffix in (".xlsx", ".xls"):
        return chunk_excel(path, source_file=source_file)
    elif suffix in (".txt", ".text", ".csv"):
        text = path.read_text(encoding="utf-8")
        return chunk_text(text, source_file=source_file)
    else:
        try:
            text = path.read_text(encoding="utf-8")
            return chunk_text(text, source_file=source_file)
        except Exception:
            return []


def _find_group_column(headers: list[str]) -> int | None:
    """Heuristic: find a column likely to be a group key (e.g. table name)."""
    group_keywords = ("表名", "table_name", "模型名称", "sheet", "分类", "类别")
    for i, h in enumerate(headers):
        hl = h.lower()
        for kw in group_keywords:
            if kw.lower() in hl:
                return i
    return None


def _rows_to_text(headers: list[str], rows: list[list[Any]]) -> str:
    """Convert rows to a readable text block."""
    lines: list[str] = []
    for row in rows:
        parts = []
        for i, v in enumerate(row):
            if v is None or str(v).strip() == "":
                continue
            h = headers[i] if i < len(headers) else f"COL_{i}"
            parts.append(f"{h}: {v}")
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines)
